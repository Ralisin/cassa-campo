from datetime import datetime
from decimal import Decimal
from io import BytesIO

from fastapi import APIRouter
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy import func, select

from app.dependencies import CurrentCassa, DbSession, OperatorMembership
from app.models import (
    BalanceType,
    Cassa,
    CassaKind,
    Movement,
    MovementReceipt,
    MovementReimbursement,
    MovementType,
    PaymentMethod,
    TransferType,
    TreasuryTransfer,
)
from app.routers.settings import latest_settings

router = APIRouter(prefix="/exports", tags=["exports"])

BLUE = "9FC5E8"
PALE_BLUE = "D9EAF7"
PALE_GREEN = "D9EAD3"
YELLOW = "FFF200"
BLACK = "000000"
WHITE = "FFFFFF"
EURO = '€ #,##0.00;[Red]-€ #,##0.00;€ 0.00'
DATE = "d/m/yyyy"
VALIDATION_LAST_ROW = 500


def _apply_border(cell, *, dotted: bool = False) -> None:
    style = "dotted" if dotted else "thin"
    cell.border = Border(
        left=Side(style=style, color=BLACK),
        right=Side(style=style, color=BLACK),
        top=Side(style=style, color=BLACK),
        bottom=Side(style=style, color=BLACK),
    )


def _merge_label(sheet, range_ref: str, value: str, *, fill: str | None = None) -> None:
    sheet.merge_cells(range_ref)
    cell = sheet[range_ref.split(":")[0]]
    cell.value = value
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)
    for row in sheet[range_ref]:
        for item in row:
            _apply_border(item)


def _add_dropdown(sheet, range_ref: str, values: list[str]) -> None:
    formula = f'"{",".join(values)}"'
    validation = DataValidation(
        type="list",
        formula1=formula,
        allow_blank=True,
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="Valore non valido",
        error="Scegli un valore dalla lista.",
    )
    sheet.add_data_validation(validation)
    validation.add(range_ref)


def _add_checks_sheet(workbook: Workbook, db: DbSession, cassa: Cassa) -> None:
    sheet = workbook.create_sheet("Controlli")
    sheet.append(["Controllo", "Valore", "Nota"])
    for cell in sheet[1]:
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor="12372A")
    missing_receipts = db.scalar(
        select(func.count(Movement.id))
        .outerjoin(Movement.receipts)
        .where(
            Movement.cassa_id == cassa.id,
            Movement.deleted_at.is_(None),
            Movement.type == MovementType.EXPENSE,
            MovementReceipt.id.is_(None),
        )
    ) or 0
    pending_reimbursements = db.scalar(
        select(func.coalesce(func.sum(Movement.amount), 0))
        .join(Movement.reimbursement)
        .where(
            Movement.cassa_id == cassa.id,
            Movement.deleted_at.is_(None),
            MovementReimbursement.reimbursed_at.is_(None),
        )
    ) or Decimal("0")
    sheet.append(["Uscite senza scontrino", missing_receipts, "Da verificare prima della chiusura"])
    sheet.append(["Rimborsi aperti", pending_reimbursements, "Importo ancora da restituire"])
    sheet["B3"].number_format = EURO
    sheet.column_dimensions["A"].width = 26
    sheet.column_dimensions["B"].width = 16
    sheet.column_dimensions["C"].width = 42


def _movement_row(movement: Movement) -> dict:
    cash_in = cash_out = bank_in = bank_out = Decimal("0")
    if movement.payment_method == PaymentMethod.CASH:
        if movement.type == MovementType.INCOME:
            cash_in = movement.amount
        else:
            cash_out = movement.amount
    elif movement.type == MovementType.INCOME:
        bank_in = movement.amount
    else:
        bank_out = movement.amount
    return {
        "created_at": movement.created_at,
        "operation_date": movement.operation_date,
        "supplier": movement.supplier,
        "unit": movement.unit,
        "balance_type": movement.balance_type.value,
        "notes": movement.notes,
        "cash_in": cash_in,
        "cash_out": cash_out,
        "bank_in": bank_in,
        "bank_out": bank_out,
        "expense": movement.amount if movement.type == MovementType.EXPENSE else Decimal("0"),
    }


def _transfer_row(transfer: TreasuryTransfer, unit: str) -> dict:
    withdrawal = transfer.type == TransferType.WITHDRAWAL
    return {
        "created_at": transfer.created_at,
        "operation_date": transfer.operation_date,
        "supplier": "Giroconto",
        "unit": unit,
        "balance_type": BalanceType.CAMP.value,
        "notes": transfer.notes,
        "cash_in": transfer.amount if withdrawal else Decimal("0"),
        "cash_out": transfer.amount if not withdrawal else Decimal("0"),
        "bank_in": transfer.amount if not withdrawal else Decimal("0"),
        "bank_out": transfer.amount if withdrawal else Decimal("0"),
        "expense": Decimal("0"),
    }


def build_excel_report(db: DbSession, cassa: Cassa) -> Workbook:
    settings = latest_settings(db, cassa.id)
    participants = settings.participants if settings else 0
    quota = settings.quota_per_person if settings else Decimal("0")
    cash_initial = settings.cash_initial if settings else Decimal("0")
    entries = [
        *[
            _movement_row(item)
            for item in db.scalars(
                select(Movement).where(
                    Movement.cassa_id == cassa.id,
                    Movement.deleted_at.is_(None),
                )
            ).all()
        ],
        *[
            _transfer_row(item, cassa.unit)
            for item in db.scalars(
                select(TreasuryTransfer).where(TreasuryTransfer.cassa_id == cassa.id)
            ).all()
        ],
    ]
    entries.sort(key=lambda item: (item["operation_date"], item["created_at"]))

    workbook = Workbook()
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    sheet = workbook.active
    sheet.title = f"Bilancio Campo {settings.camp_year if settings else datetime.now().year}"
    sheet.freeze_panes = "A8"
    sheet.sheet_view.showGridLines = True
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.print_title_rows = "1:7"

    widths = {
        "A": 6,
        "B": 13,
        "C": 13,
        "D": 24,
        "E": 9,
        "F": 10,
        "G": 38,
        "H": 3,
        "I": 3,
        "J": 13,
        "K": 13,
        "L": 14,
        "M": 13,
        "N": 13,
        "O": 14,
        "P": 14,
        "Q": 16,
        "R": 16,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    sheet.column_dimensions["H"].hidden = True
    sheet.column_dimensions["I"].hidden = True

    year = settings.camp_year if settings else datetime.now().year
    camp_name = settings.camp_name if settings else "Campo"
    _merge_label(sheet, "A1:C3", f"{year - 1}-{year}")
    sheet["G1"] = f"{camp_name} {year}"
    sheet["G1"].font = Font(bold=True, size=12)
    sheet["D2"] = "Numero Paganti"
    sheet["E2"] = participants
    sheet["D3"] = "Quota totale a ragazzo"
    sheet["E3"] = float(quota)
    for coordinate in ("D2", "D3"):
        sheet[coordinate].font = Font(bold=True)
        sheet[coordinate].alignment = Alignment(horizontal="right")
    for coordinate in ("E2", "E3"):
        sheet[coordinate].fill = PatternFill("solid", fgColor=YELLOW)
        sheet[coordinate].font = Font(bold=True)
        _apply_border(sheet[coordinate])
    sheet["E3"].number_format = EURO

    _merge_label(sheet, "J2:K2", "Spesa Massima")
    sheet["L2"] = "=E2*E3"
    _merge_label(sheet, "M2:N2", "Spesa Eseguita")
    sheet["O2"] = f"=SUM(K8:K{max(8, 7 + len(entries))})+SUM(N8:N{max(8, 7 + len(entries))})"
    _merge_label(sheet, "Q2:R2", "Cassa residua")
    sheet.merge_cells("Q3:R4")
    sheet["Q3"] = "=L2-O2"
    for coordinate in ("L2", "O2", "Q3"):
        sheet[coordinate].font = Font(bold=True, size=12)
        sheet[coordinate].alignment = Alignment(horizontal="center", vertical="center")
        sheet[coordinate].number_format = EURO
        _apply_border(sheet[coordinate])

    sheet.merge_cells("A4:C5")
    sheet["A4"] = datetime.now().date()
    sheet["A4"].number_format = DATE
    sheet["A4"].font = Font(bold=True, italic=True)
    sheet["A4"].alignment = Alignment(horizontal="center", vertical="center")

    sheet.merge_cells("E4:E7")
    sheet.merge_cells("F4:F7")
    sheet.merge_cells("G4:I7")
    sheet["G4"] = "Note"
    sheet.merge_cells("J4:L5")
    sheet["J4"] = "Cassa Contanti Branca"
    sheet.merge_cells("M4:O5")
    sheet["M4"] = "Cassa Banca Branca"
    sheet.merge_cells("P4:P7")
    for coordinate in ("E4", "F4", "G4", "J4", "M4", "P4"):
        sheet[coordinate].alignment = Alignment(horizontal="center", vertical="center")
        sheet[coordinate].font = Font(bold=True)

    headers = {
        "A6": "N.°",
        "B6": "Data reg.ne",
        "C6": "Data oper.ne",
        "D6": "Fornitore",
        "E4": "Unità",
        "F4": "Bilancio",
        "G4": "Note",
        "J6": "Entrate",
        "K6": "Uscite",
        "L6": "Saldo",
        "M6": "Entrate",
        "N6": "Uscite",
        "O6": "Saldo",
        "P4": "Parziale EG",
    }
    sheet.merge_cells("A6:A7")
    sheet.merge_cells("B6:B7")
    sheet.merge_cells("C6:C7")
    sheet.merge_cells("D4:D5")
    sheet.merge_cells("D6:D7")
    for coordinate, value in headers.items():
        sheet[coordinate] = value
    for row in sheet.iter_rows(min_row=4, max_row=7, min_col=1, max_col=16):
        for cell in row:
            if cell.column not in (8, 9):
                cell.fill = PatternFill("solid", fgColor=BLUE)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                _apply_border(cell)
    sheet["L7"] = float(cash_initial)
    sheet["O7"] = "=L2-L7"
    for coordinate in ("L7", "O7"):
        sheet[coordinate].fill = PatternFill("solid", fgColor=PALE_GREEN)
        sheet[coordinate].font = Font(bold=True)
        sheet[coordinate].number_format = EURO
        _apply_border(sheet[coordinate])

    for index, entry in enumerate(entries, start=1):
        row = index + 7
        values = {
            "A": index,
            "B": entry["created_at"].date(),
            "C": entry["operation_date"],
            "D": entry["supplier"],
            "E": entry["unit"],
            "F": entry["balance_type"],
            "G": entry["notes"],
            "J": float(entry["cash_in"]) or None,
            "K": float(entry["cash_out"]) or None,
            "M": float(entry["bank_in"]) or None,
            "N": float(entry["bank_out"]) or None,
        }
        for column, value in values.items():
            sheet[f"{column}{row}"] = value
        previous = row - 1
        sheet[f"L{row}"] = f"=L{previous}+J{row}-K{row}"
        sheet[f"O{row}"] = f"=O{previous}+M{row}-N{row}"
        sheet[f"P{row}"] = f"=J{row}-K{row}+M{row}-N{row}"
        for column in range(1, 17):
            cell = sheet.cell(row, column)
            if column not in (8, 9):
                _apply_border(cell, dotted=True)
                if column in (1, 5, 6):
                    cell.alignment = Alignment(horizontal="center")
                if column in (10, 11, 12, 13, 14, 15, 16):
                    cell.number_format = EURO
        for column in ("B", "C"):
            sheet[f"{column}{row}"].number_format = DATE
        for column in ("L", "O"):
            sheet[f"{column}{row}"].fill = PatternFill("solid", fgColor=PALE_GREEN)

    last_row = max(8, 7 + len(entries))
    sheet.auto_filter.ref = f"A7:P{last_row}"
    sheet.print_area = f"A1:R{last_row}"
    for column in range(1, 19):
        sheet.cell(7, column).font = Font(bold=True)
    _add_dropdown(sheet, f"E8:E{VALIDATION_LAST_ROW}", ["L/C", "E/G", "R/S", "CoCa", "Gruppo"])
    _add_dropdown(sheet, f"F8:F{VALIDATION_LAST_ROW}", ["O", "C", "A"])
    _add_checks_sheet(workbook, db, cassa)
    return workbook


def build_annual_excel_report(db: DbSession, cassa: Cassa) -> Workbook:
    entries = [
        _movement_row(item)
        for item in db.scalars(
            select(Movement).where(
                Movement.cassa_id == cassa.id,
                Movement.deleted_at.is_(None),
            )
        ).all()
    ]
    entries.sort(key=lambda item: (item["operation_date"], item["created_at"]))

    workbook = Workbook()
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    sheet = workbook.active
    sheet.title = f"Cassa Anno {cassa.year}"
    sheet.freeze_panes = "A5"
    sheet.sheet_view.showGridLines = True

    headers = [
        "N.°",
        "Data reg.ne",
        "Data oper.ne",
        "Fornitore",
        "Unità",
        "Bilancio",
        "Note",
        "Entrate contanti",
        "Uscite contanti",
        "Entrate banca",
        "Uscite banca",
        "Saldo",
    ]
    widths = [6, 13, 13, 28, 10, 10, 38, 15, 15, 15, 15, 15]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[sheet.cell(1, index).column_letter].width = width

    sheet.merge_cells("A1:L1")
    sheet["A1"] = f"Cassa Anno {cassa.unit} {cassa.year}"
    sheet["A1"].font = Font(bold=True, size=14)
    sheet["A1"].alignment = Alignment(horizontal="center")
    sheet["A2"] = "Aperta"
    sheet["B2"] = cassa.opened_at
    sheet["D2"] = "Chiusa"
    sheet["E2"] = cassa.closed_at
    for coordinate in ("B2", "E2"):
        sheet[coordinate].number_format = DATE

    for column, value in enumerate(headers, start=1):
        cell = sheet.cell(4, column)
        cell.value = value
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        _apply_border(cell)

    for index, entry in enumerate(entries, start=1):
        row = index + 4
        values = [
            index,
            entry["created_at"].date(),
            entry["operation_date"],
            entry["supplier"],
            entry["unit"],
            entry["balance_type"],
            entry["notes"],
            float(entry["cash_in"]) or None,
            float(entry["cash_out"]) or None,
            float(entry["bank_in"]) or None,
            float(entry["bank_out"]) or None,
        ]
        for column, value in enumerate(values, start=1):
            sheet.cell(row, column).value = value
        sheet.cell(row, 12).value = f"=H{row}-I{row}+J{row}-K{row}"
        for column in range(1, 13):
            cell = sheet.cell(row, column)
            _apply_border(cell, dotted=True)
            if column in (2, 3):
                cell.number_format = DATE
            if column >= 8:
                cell.number_format = EURO
    last_row = max(5, 4 + len(entries))
    sheet.auto_filter.ref = f"A4:L{last_row}"
    sheet.print_area = f"A1:L{last_row}"
    _add_checks_sheet(workbook, db, cassa)
    return workbook


def build_cassa_excel_report(db: DbSession, cassa: Cassa) -> Workbook:
    if cassa.kind == CassaKind.ANNO:
        return build_annual_excel_report(db, cassa)
    return build_excel_report(db, cassa)


@router.get("/excel")
def export_excel(db: DbSession, cassa: CurrentCassa, _: OperatorMembership) -> StreamingResponse:
    workbook = build_cassa_excel_report(db, cassa)
    filename = "bilancio-anno.xlsx" if cassa.kind == CassaKind.ANNO else "bilancio-campo.xlsx"
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
